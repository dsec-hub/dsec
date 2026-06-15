"""Parse the two weekly DUSA workbooks into plain dicts.

Kept deliberately free of DB/HTTP concerns so it is trivially unit-testable
against the real sample files in ``tests/fixtures``.

Two shapes:

* **Membership** — a single ``Report`` sheet, header on row 1, one member per
  row thereafter.
* **P&L** — a multi-sheet workbook; we read the ``Detailed Club Transactions``
  tab (a clean ledger) for transactions and derive the headline balances from
  it (income negative, expense positive, reserve/opening in the 3xxx accounts).
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from typing import Any

from openpyxl import load_workbook

# Excel's 1900 date system counts days from this epoch (accounts for the
# historical 1900-leap-year bug by using Dec 30 1899).
_EXCEL_EPOCH = datetime(1899, 12, 30)


def excel_to_date(value: Any) -> date | None:
    """Coerce an Excel cell (serial number, datetime, or date string) to a date."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return (_EXCEL_EPOCH + timedelta(days=float(value))).date()
    # Fall back to a few common string formats.
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _num(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


# ----------------------------------------------------------------------------
# Membership report
# ----------------------------------------------------------------------------

# Maps spreadsheet headers (lower-cased, stripped) to our field names. Robust to
# minor header wording changes by matching on a substring.
_MEMBER_HEADER_MAP = [
    ("full name", "full_name"),
    ("student id", "student_id"),
    ("email", "email"),
    ("campus", "campus"),
    ("first subscription", "first_subscription_date"),
    ("last paid", "last_paid_date"),
    ("end date", "end_date"),
    ("account name", "faculty"),
    ("new/renewal", "membership_type"),
    ("payment option", "payment_option"),
]

_DATE_FIELDS = {"first_subscription_date", "last_paid_date", "end_date"}


@dataclass
class MembershipParse:
    members: list[dict] = field(default_factory=list)
    total: int = 0
    dusa_member_count: int = 0
    non_dusa_count: int = 0
    new_count: int = 0
    renewal_count: int = 0


def parse_membership(data: bytes) -> MembershipParse:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb["Report"] if "Report" in wb.sheetnames else wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return MembershipParse()

    # Build a column-index -> field mapping from the header row.
    col_field: dict[int, str] = {}
    for idx, cell in enumerate(header):
        label = str(cell or "").strip().lower()
        for needle, field_name in _MEMBER_HEADER_MAP:
            if needle in label:
                col_field[idx] = field_name
                break

    out = MembershipParse()
    for row in rows:
        rec: dict[str, Any] = {}
        for idx, field_name in col_field.items():
            val = row[idx] if idx < len(row) else None
            if field_name in _DATE_FIELDS:
                rec[field_name] = excel_to_date(val)
            elif field_name == "student_id":
                rec[field_name] = _clean_student_id(val)
            else:
                rec[field_name] = (str(val).strip() if val not in (None, "") else None)
        if not rec.get("student_id") and not rec.get("full_name"):
            continue  # skip blank/trailing rows
        rec["dusa_member"] = _is_dusa_member(rec.get("payment_option"))
        out.members.append(rec)

    wb.close()

    out.total = len(out.members)
    out.dusa_member_count = sum(1 for m in out.members if m.get("dusa_member"))
    out.non_dusa_count = out.total - out.dusa_member_count
    out.new_count = sum(
        1 for m in out.members if (m.get("membership_type") or "").lower() == "new"
    )
    out.renewal_count = sum(
        1 for m in out.members if (m.get("membership_type") or "").lower() == "renewal"
    )
    return out


def _clean_student_id(val: Any) -> str | None:
    if val is None or val == "":
        return None
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val).strip()


def _is_dusa_member(payment_option: str | None) -> bool:
    # e.g. "BW Software Eng (DUSA Member)" vs "... (Non-DUSA Member)".
    if not payment_option:
        return False
    p = payment_option.lower()
    return "dusa member" in p and "non-dusa" not in p and "non dusa" not in p


# ----------------------------------------------------------------------------
# P&L report
# ----------------------------------------------------------------------------

_TX_HEADERS = {
    "posting date": "posting_date",
    "document no.": "document_no",
    "document no": "document_no",
    "g/l account no.": "gl_account_no",
    "g/l account no": "gl_account_no",
    "g/l account name": "gl_account_name",
    "description": "description",
    "department code": "department_code",
    "clubs code": "club_code",
    "amount": "amount",
}


@dataclass
class PnLParse:
    transactions: list[dict] = field(default_factory=list)
    opening_balance: float = 0.0
    total_income: float = 0.0
    total_expense: float = 0.0
    closing_balance: float = 0.0
    fy_start: date | None = None


def parse_pnl(data: bytes) -> PnLParse:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)

    out = PnLParse()
    out.fy_start = _read_fy_start(wb)

    sheet_name = _find_sheet(wb, "Detailed Club Transactions")
    if sheet_name is None:
        wb.close()
        return out
    ws = wb[sheet_name]

    # Locate the header row (it isn't row 1 — there are filter/label rows above).
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    header_row_idx, col_field = _locate_tx_header(grid)
    if header_row_idx is None:
        wb.close()
        return out

    for row in grid[header_row_idx + 1:]:
        rec = _parse_tx_row(row, col_field)
        if rec is None:
            continue
        out.transactions.append(rec)

    wb.close()

    # Derive headline balances from the ledger (robust vs the messy pivot tab).
    # Sign convention: income negative, expense positive, reserve/opening 3xxx.
    out.opening_balance = round(
        -sum(t["amount"] for t in out.transactions if _acct_kind(t["gl_account_no"]) == "balance"), 2
    )
    out.total_income = round(
        -sum(t["amount"] for t in out.transactions if _acct_kind(t["gl_account_no"]) == "income"), 2
    )
    out.total_expense = round(
        sum(t["amount"] for t in out.transactions if _acct_kind(t["gl_account_no"]) == "expense"), 2
    )
    out.closing_balance = round(-sum(t["amount"] for t in out.transactions), 2)
    return out


def _parse_tx_row(row: list, col_field: dict[int, str]) -> dict | None:
    rec: dict[str, Any] = {}
    for idx, field_name in col_field.items():
        val = row[idx] if idx < len(row) else None
        if field_name == "posting_date":
            rec[field_name] = excel_to_date(val)
        elif field_name == "amount":
            rec[field_name] = _num(val)
        elif field_name == "gl_account_no":
            rec[field_name] = _clean_student_id(val)  # same int->str cleanup
        else:
            rec[field_name] = (str(val).strip() if val not in (None, "") else None)
    # Require at least an account number and an amount to count as a real line.
    if rec.get("gl_account_no") is None or rec.get("amount") is None:
        return None
    rec.setdefault("amount", 0.0)
    rec["amount"] = rec["amount"] or 0.0
    rec["amount_abs"] = round(abs(rec["amount"]), 2)
    rec["kind"] = _acct_kind(rec["gl_account_no"])
    return rec


def _acct_kind(gl_account_no: str | None) -> str:
    """DUSA chart of accounts: 3xxx balance/reserve, 4xxx & 8xxx income, 6xxx expense."""
    if not gl_account_no:
        return "other"
    head = str(gl_account_no).strip()[:1]
    if head == "3":
        return "balance"
    if head in ("4", "8"):
        return "income"
    if head == "6":
        return "expense"
    return "other"


def _find_sheet(wb, name: str) -> str | None:
    for s in wb.sheetnames:
        if s.strip().lower() == name.lower():
            return s
    return None


def _locate_tx_header(grid: list[list]) -> tuple[int | None, dict[int, str]]:
    """Find the row that holds the transaction headers; return (row_idx, mapping).

    The workbook also has hidden ``Headers:``/``Fields:`` metadata rows that
    repeat the column names at *shifted* positions, so a header-text match alone
    is ambiguous. We disambiguate by requiring the rows *beneath* a candidate to
    actually parse as transactions — only the real header sits above the ledger.
    """
    for i, row in enumerate(grid[:40]):
        mapping: dict[int, str] = {}
        for idx, cell in enumerate(row):
            label = str(cell or "").strip().lower()
            if label in _TX_HEADERS:
                mapping[idx] = _TX_HEADERS[label]
        fields = set(mapping.values())
        if not {"posting_date", "amount", "gl_account_no"}.issubset(fields):
            continue
        # Validate: at least one of the next few rows must be a real ledger line.
        if any(_parse_tx_row(r, mapping) is not None for r in grid[i + 1:i + 6]):
            return i, mapping
    return None, {}


def _read_fy_start(wb) -> date | None:
    name = _find_sheet(wb, "Options")
    if name is None:
        return None
    ws = wb[name]
    for row in ws.iter_rows(values_only=True):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        if "fy start date" in cells:
            pos = cells.index("fy start date")
            if pos + 1 < len(row):
                return excel_to_date(row[pos + 1])
    return None
