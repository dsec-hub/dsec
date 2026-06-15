"""Tests for the DUSA weekly-import ingestion (parser + /ingest/dusa endpoint).

Fixtures are built in-memory with openpyxl to mirror the *real* DUSA workbooks
(including the P&L tab's hidden ``Headers:``/``Fields:`` decoy rows) — no real
member PII is committed to the repo.
"""

from __future__ import annotations

import io
from datetime import date

import pytest
from openpyxl import Workbook

from app import models
from app.core.apikeys import generate_key
from app.features.ingest.parser import excel_to_date, parse_membership, parse_pnl

_EPOCH = date(1899, 12, 30)


def _serial(d: date) -> int:
    return (d - _EPOCH).days


# ---------------------------------------------------------------------------
# Synthetic workbook builders (mirror the real structures)
# ---------------------------------------------------------------------------

_MEMBER_HEADERS = [
    "Full Name", "Student ID", "Email", "Campus",
    "First Subscription Date", "Last Paid Date", "End Date",
    "Account Name: Account Name", "New/Renewal", "Payment Option Name",
]

# (name, student_id, email, campus, first, last, end, faculty, new/renewal, payment)
_SAMPLE_MEMBERS = [
    ("Alex EXAMPLE", 200000001, "a@x.edu", "Burwood",
     date(2025, 6, 26), date(2025, 6, 26), date(2026, 6, 26),
     "Science, Engineering and Built Environment", "New", "BW Software Eng (DUSA Member)"),
    ("Blair SAMPLE", 200000002, "b@x.edu", "Burwood",
     date(2025, 6, 30), date(2025, 6, 30), date(2026, 6, 30),
     "Business and Law", "New", "BW Software Eng (Non-DUSA Member)"),
    ("Casey TESTER", 200000003, "c@x.edu", "Burwood",
     date(2024, 7, 1), date(2025, 7, 1), date(2026, 7, 1),
     "Science, Engineering and Built Environment", "Renewal", "BW Software Eng (DUSA Member)"),
]


def build_membership_xlsx(members=_SAMPLE_MEMBERS) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.append(_MEMBER_HEADERS)
    for m in members:
        name, sid, email, campus, first, last, end, faculty, kind, pay = m
        # Dates as Excel serial numbers, exactly like the real DUSA export.
        ws.append([name, sid, email, campus, _serial(first), _serial(last),
                   _serial(end), faculty, kind, pay])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# (posting_date, document_no, gl_no, gl_name, description, dept, club, amount)
_SAMPLE_TX = [
    (date(2026, 1, 1), "GJ-3000654", 3005, "Club Reserves", "Rollover", 100, "BW DSEC", -1970.09),
    (date(2026, 3, 13), "GJ-4006520", 4005, "Grants Received", "T1 Grant", 210, "BW DSEC", -75.00),
    (date(2026, 1, 6), "SF003914", 4006, "Memberships - Clubs", "Card", 210, "BW DSEC", -100.00),
    (date(2026, 3, 3), "PPI-011494", 6541, "Food and Beverage Expenses", "Pizza", 210, "BW DSEC", 113.98),
    (date(2026, 4, 1), "PPI-011626", 6543, "Venue Hire", "Venue", 210, "BW DSEC", 462.73),
]


def build_pnl_xlsx(transactions=_SAMPLE_TX, fy_start=date(2026, 1, 1)) -> bytes:
    wb = Workbook()
    # Options sheet with the FY start date.
    opts = wb.active
    opts.title = "Options"
    opts.append(["Option", "FY Start Date", _serial(fy_start)])

    ws = wb.create_sheet("Detailed Club Transactions")
    tx_cols = ["Posting Date", "Document No.", "G/L Account No.", "G/L Account Name",
               "Description", "Department Code", "Clubs Code", "Amount"]

    # Explicit row counter — openpyxl's `append([])` writes no cells and does not
    # advance max_row, so we place cells by absolute (row, column) instead.
    r = 1
    ws.cell(row=r, column=1, value="filler"); r += 1

    # Decoy metadata rows: column names live at a SHIFTED position (col 12+),
    # exactly like the real workbook — the parser must not latch onto these.
    for label in ("Headers:", "Fields:"):
        ws.cell(row=r, column=4, value=label)
        for i, name in enumerate(tx_cols):
            ws.cell(row=r, column=12 + i, value=name)
        r += 1

    # The REAL header row: column names at col 4..11 (0-indexed 3..10).
    for i, name in enumerate(tx_cols):
        ws.cell(row=r, column=4 + i, value=name)
    r += 1

    # Data rows aligned under the real header.
    for tx in transactions:
        pdate, doc, gl, glname, desc, dept, club, amt = tx
        vals = [_serial(pdate), doc, gl, glname, desc, dept, club, amt]
        for i, v in enumerate(vals):
            ws.cell(row=r, column=4 + i, value=v)
        r += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Parser unit tests
# ---------------------------------------------------------------------------

def test_excel_to_date_variants():
    assert excel_to_date(_serial(date(2026, 6, 12))) == date(2026, 6, 12)
    assert excel_to_date(date(2026, 1, 1)) == date(2026, 1, 1)
    assert excel_to_date("2026-06-12") == date(2026, 6, 12)
    assert excel_to_date("12/06/2026") == date(2026, 6, 12)
    assert excel_to_date(None) is None
    assert excel_to_date("") is None


def test_parse_membership_counts_and_dates():
    p = parse_membership(build_membership_xlsx())
    assert p.total == 3
    assert p.dusa_member_count == 2
    assert p.non_dusa_count == 1
    assert p.new_count == 2
    assert p.renewal_count == 1
    first = p.members[0]
    assert first["student_id"] == "200000001"  # numeric id coerced to clean str
    assert first["end_date"] == date(2026, 6, 26)
    assert first["dusa_member"] is True
    assert p.members[1]["dusa_member"] is False  # Non-DUSA Member


def test_parse_pnl_skips_decoy_headers_and_balances():
    p = parse_pnl(build_pnl_xlsx())
    assert len(p.transactions) == 5  # decoy/metadata rows excluded
    assert p.fy_start == date(2026, 1, 1)
    assert p.opening_balance == 1970.09
    assert p.total_income == 175.00       # -(−75 − 100)
    assert p.total_expense == 576.71      # 113.98 + 462.73
    assert p.closing_balance == 1568.38   # opening + income − expense
    assert p.closing_balance == round(p.opening_balance + p.total_income - p.total_expense, 2)
    # sign convention + derived fields preserved
    income = next(t for t in p.transactions if t["gl_account_no"] == "4006")
    assert income["amount"] == -100.00 and income["kind"] == "income" and income["amount_abs"] == 100.00


# ---------------------------------------------------------------------------
# Endpoint tests
# ---------------------------------------------------------------------------

@pytest.fixture
def ingest_key(db):
    gen = generate_key()
    db.add(models.APIKey(name="ingest", prefix=gen.prefix, key_hash=gen.key_hash,
                         scopes=["read", "ingest"]))
    db.commit()
    return gen.raw_key


@pytest.fixture
def read_only_key(db):
    gen = generate_key()
    db.add(models.APIKey(name="ro", prefix=gen.prefix, key_hash=gen.key_hash, scopes=["read"]))
    db.commit()
    return gen.raw_key


def _post(client, key, *, report_type, message_id, data, filename="r.xlsx"):
    return client.post(
        "/ingest/dusa",
        headers={"Authorization": f"Bearer {key}"},
        data={"report_type": report_type, "message_id": message_id,
              "received_at": "2026-06-12T07:20:00Z", "sender": "DUSA", "subject": "weekly"},
        files={"file": (filename, data,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )


def test_ingest_requires_auth(client):
    r = client.post("/ingest/dusa", data={"report_type": "membership", "message_id": "x"},
                    files={"file": ("r.xlsx", build_membership_xlsx(), "application/octet-stream")})
    assert r.status_code == 401


def test_ingest_requires_ingest_scope(client, read_only_key):
    r = _post(client, read_only_key, report_type="membership", message_id="m1",
              data=build_membership_xlsx())
    assert r.status_code == 403


def test_ingest_membership_lands_rows(client, ingest_key, db):
    r = _post(client, ingest_key, report_type="membership", message_id="mb-1",
              data=build_membership_xlsx())
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["status"] == "ingested"
    assert body["rows_ingested"] == 3
    assert body["membership"]["dusa_member_count"] == 2

    assert db.query(models.Member).count() == 3
    assert db.query(models.Member).filter_by(is_current=True).count() == 3
    assert db.query(models.MemberReport).count() == 1


def test_ingest_membership_marks_dropped_members_not_current(client, ingest_key, db):
    _post(client, ingest_key, report_type="membership", message_id="wk1",
          data=build_membership_xlsx())
    # Next week only the first member remains.
    _post(client, ingest_key, report_type="membership", message_id="wk2",
          data=build_membership_xlsx(members=_SAMPLE_MEMBERS[:1]))
    assert db.query(models.Member).count() == 3            # roster retained
    assert db.query(models.Member).filter_by(is_current=True).count() == 1  # only the renewer


def test_ingest_pnl_lands_transactions(client, ingest_key, db):
    r = _post(client, ingest_key, report_type="pnl", message_id="pl-1", data=build_pnl_xlsx())
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["rows_ingested"] == 5
    assert body["finance"]["closing_balance"] == 1568.38

    report = db.query(models.FinanceReport).filter_by(is_current=True).one()
    assert report.transaction_count == 5
    assert db.query(models.FinanceTransaction).filter_by(report_id=report.id).count() == 5


def test_ingest_is_idempotent(client, ingest_key):
    first = _post(client, ingest_key, report_type="membership", message_id="dup-1",
                  data=build_membership_xlsx())
    assert first.status_code == 200
    again = _post(client, ingest_key, report_type="membership", message_id="dup-1",
                  data=build_membership_xlsx())
    assert again.status_code == 409


def test_ingest_rejects_unknown_report_type(client, ingest_key):
    r = _post(client, ingest_key, report_type="nonsense", message_id="z1",
              data=build_membership_xlsx())
    assert r.status_code == 422


def test_imports_listing(client, ingest_key):
    _post(client, ingest_key, report_type="membership", message_id="li-1",
          data=build_membership_xlsx())
    r = client.get("/ingest/imports", headers={"Authorization": f"Bearer {ingest_key}"})
    assert r.status_code == 200
    rows = r.json()
    assert rows and rows[0]["message_id"] == "li-1" and rows[0]["status"] == "ok"
